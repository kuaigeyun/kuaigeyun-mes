"""
往来对账单业务服务
"""

from __future__ import annotations

from calendar import monthrange
from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from tortoise.expressions import Q
from tortoise import timezone as tortoise_timezone

from apps.common.audit_actor import apply_update_audit
from apps.common.base_service import AppBaseService
from apps.kuaicaiwu.constants.finance_source_types import (
    PAYABLE_SOURCE_OUTSOURCE_RECEIPT,
    PAYABLE_SOURCE_PURCHASE_RECEIPT,
    PAYABLE_SOURCE_PURCHASE_RETURN,
    RECEIVABLE_SOURCE_SALES_RETURN,
)
from apps.kuaicaiwu.models.partner_statement import PartnerStatement
from apps.kuaicaiwu.models.payable import Payable
from apps.kuaicaiwu.models.payment import Payment
from apps.kuaicaiwu.models.receivable import Receivable
from apps.kuaicaiwu.models.receipt import Receipt
from apps.master_data.models.customer import Customer
from apps.master_data.models.supplier import Supplier
from core.utils.timezone_utils import to_api_isoformat, today_site_str
from infra.exceptions.exceptions import BusinessLogicError, NotFoundError, ValidationError
from infra.models.tenant import Tenant
from infra.models.user import User

_APPROVED_REVIEW = ("已审核",)
_EXCLUDED_REVIEW = ("待审核", "驳回")
_MONEY = Decimal("0.01")

# 对账单行 doc_type → 稳定族，用于「已纳入其它对账单」去重
_DOC_TYPE_FAMILY: Dict[str, str] = {
    "应收单": "receivable",
    "销售退货": "receivable",
    "收款单": "receipt",
    "收款退款": "receipt",
    "应付单": "payable",
    "采购退货": "payable",
    "付款单": "payment",
    "付款退款": "payment",
}


def _q_money(value: Decimal | float | int | str) -> Decimal:
    return Decimal(str(value or 0)).quantize(_MONEY)


def _is_refund_voucher(settlement_type: Optional[str], total_amount: Decimal) -> bool:
    return (settlement_type or "normal") == "refund" or total_amount < 0


def _line_doc_key(doc_type: Any, doc_id: Any) -> Optional[Tuple[str, int]]:
    if doc_id is None:
        return None
    family = _DOC_TYPE_FAMILY.get(str(doc_type or "").strip())
    if not family:
        return None
    try:
        return family, int(doc_id)
    except (TypeError, ValueError):
        return None


def _abs_money(value: Decimal | float | int | str) -> Decimal:
    return abs(_q_money(value))


def _line_doc_amount(ln: Dict[str, Any]) -> Decimal:
    if ln.get("doc_amount") is not None:
        return _abs_money(ln["doc_amount"])
    debit = _q_money(ln.get("debit", 0))
    credit = _q_money(ln.get("credit", 0))
    return debit if debit > 0 else credit


def _line_statement_amount(ln: Dict[str, Any]) -> Decimal:
    if ln.get("statement_amount") is not None:
        return _q_money(ln["statement_amount"])
    return _line_doc_amount(ln)


def _line_is_debit_side(ln: Dict[str, Any]) -> bool:
    return _q_money(ln.get("debit", 0)) > 0


def _apply_statement_amount_to_line(
    ln: Dict[str, Any],
    statement_amount: Decimal,
    doc_amount: Decimal,
    prior_stated: Decimal,
    remaining: Decimal,
) -> Dict[str, Any]:
    row = dict(ln)
    row["doc_amount"] = float(doc_amount)
    row["prior_stated_amount"] = float(prior_stated)
    row["remaining_amount"] = float(remaining)
    row["statement_amount"] = float(statement_amount)
    amt = float(statement_amount)
    if _line_is_debit_side(ln):
        row["debit"] = amt
        row["credit"] = 0.0
    else:
        row["debit"] = 0.0
        row["credit"] = amt
    return row


def _normalize_statement_line_amounts(ln: Dict[str, Any]) -> Dict[str, Any]:
    """旧快照补齐单据/已对/未对/本次对账金额字段。"""
    row = dict(ln)
    doc_amount = _line_doc_amount(row)
    prior = _q_money(row.get("prior_stated_amount", 0))
    statement_amount = _line_statement_amount(row)
    remaining = _q_money(doc_amount - prior)
    return _apply_statement_amount_to_line(row, statement_amount, doc_amount, prior, remaining)


def _extract_waste_quantities(other_checks: Any) -> Tuple[Optional[float], Optional[float]]:
    """从检验 other_checks JSON 提取工废/料废数量。"""
    if not isinstance(other_checks, dict):
        return None, None

    def _pick(*keys: str) -> Optional[float]:
        for key in keys:
            val = other_checks.get(key)
            if val is not None and val != "":
                try:
                    return float(val)
                except (TypeError, ValueError):
                    continue
        return None

    process_waste = _pick("process_waste_qty", "process_waste", "工废", "工废数量")
    material_waste = _pick("material_waste_qty", "material_waste", "料废", "料废数量")
    return process_waste, material_waste


def _inspection_passed(inspection: Any) -> Optional[bool]:
    if inspection is None:
        return None
    result = str(getattr(inspection, "inspection_result", "") or "").strip()
    quality = str(getattr(inspection, "quality_status", "") or "").strip()
    if result in ("合格", "通过") or quality == "合格":
        return True
    if result in ("不合格", "未通过", "拒收") or quality == "不合格":
        return False
    return None


def period_to_date_range(period: str) -> Tuple[date, date]:
    """YYYY-MM -> (月初, 月末)"""
    try:
        year_s, month_s = period.split("-", 1)
        year, month = int(year_s), int(month_s)
    except (ValueError, AttributeError) as e:
        raise ValidationError("对账周期格式应为 YYYY-MM") from e
    if month < 1 or month > 12:
        raise ValidationError("对账周期月份无效")
    start = date(year, month, 1)
    end = date(year, month, monthrange(year, month)[1])
    return start, end


class PartnerStatementService(AppBaseService[PartnerStatement]):
    """往来对账单服务

    对账单明细行 doc_type 与前端展示说明（唯一文案来源）：
    - 应收单：销售出库/开票等正向应收
    - 收款单：客户正常收款（贷方）
    - 收款退款：settlement_type=refund 或负金额收款（借方冲回）
    - 销售退货：销售退货关联红字应收（贷方冲减）
    - 应付单：采购入库/委外等正向应付
    - 付款单：供应商正常付款（贷方）
    - 付款退款：settlement_type=refund 或负金额付款（借方冲回）
    - 采购退货：采购退货关联红字应付（贷方冲减）
    """

    async def _get_tenant_company_name(self, tenant_id: int) -> str:
        tenant = await Tenant.get_or_none(id=tenant_id)
        return (tenant.name if tenant else "") or "本公司"

    async def _load_partner_snapshot(
        self, tenant_id: int, partner_id: int, partner_type: str
    ) -> Dict[str, Any]:
        if partner_type == "Customer":
            obj = await Customer.get_or_none(tenant_id=tenant_id, id=partner_id, deleted_at__isnull=True)
            if not obj:
                raise NotFoundError(f"客户不存在: {partner_id}")
            return {
                "id": obj.id,
                "name": obj.name or str(partner_id),
                "finance_contact_name": getattr(obj, "finance_contact_name", None),
                "finance_contact_phone": getattr(obj, "finance_contact_phone", None),
                "finance_contact_email": getattr(obj, "finance_contact_email", None),
            }
        if partner_type == "Supplier":
            obj = await Supplier.get_or_none(tenant_id=tenant_id, id=partner_id, deleted_at__isnull=True)
            if not obj:
                raise NotFoundError(f"供应商不存在: {partner_id}")
            return {
                "id": obj.id,
                "name": obj.name or str(partner_id),
                "finance_contact_name": getattr(obj, "finance_contact_name", None),
                "finance_contact_phone": getattr(obj, "finance_contact_phone", None),
                "finance_contact_email": getattr(obj, "finance_contact_email", None),
            }
        raise ValidationError("partner_type 必须为 Customer 或 Supplier")

    def _approved_doc_filter(self) -> Q:
        return Q(review_status__in=_APPROVED_REVIEW) | Q(review_status__isnull=True)

    async def _sum_debits_before(
        self, tenant_id: int, partner_id: int, partner_type: str, before_date: date
    ) -> Decimal:
        total = Decimal("0")
        if partner_type == "Customer":
            rows = await Receivable.filter(
                tenant_id=tenant_id,
                customer_id=partner_id,
                business_date__lt=before_date,
                deleted_at__isnull=True,
            ).exclude(review_status__in=_EXCLUDED_REVIEW).exclude(
                source_type=RECEIVABLE_SOURCE_SALES_RETURN
            ).all()
            total += sum((r.total_amount for r in rows), Decimal("0"))

            receipts = await Receipt.filter(
                tenant_id=tenant_id,
                customer_id=partner_id,
                receipt_date__lt=before_date,
                status="Confirmed",
                deleted_at__isnull=True,
            ).all()
            for r in receipts:
                amt = _q_money(r.total_amount)
                if _is_refund_voucher(getattr(r, "settlement_type", None), amt):
                    total += _abs_money(amt)
        else:
            rows = await Payable.filter(
                tenant_id=tenant_id,
                supplier_id=partner_id,
                business_date__lt=before_date,
                deleted_at__isnull=True,
            ).exclude(review_status__in=_EXCLUDED_REVIEW).exclude(
                source_type=PAYABLE_SOURCE_PURCHASE_RETURN
            ).all()
            total += sum((p.total_amount for p in rows), Decimal("0"))

            payments = await Payment.filter(
                tenant_id=tenant_id,
                supplier_id=partner_id,
                payment_date__lt=before_date,
                status="Confirmed",
                deleted_at__isnull=True,
            ).all()
            for p in payments:
                amt = _q_money(p.total_amount)
                if _is_refund_voucher(getattr(p, "settlement_type", None), amt):
                    total += _abs_money(amt)
        return _q_money(total)

    async def _sum_credits_before(
        self, tenant_id: int, partner_id: int, partner_type: str, before_date: date
    ) -> Decimal:
        total = Decimal("0")
        if partner_type == "Customer":
            rows = await Receipt.filter(
                tenant_id=tenant_id,
                customer_id=partner_id,
                receipt_date__lt=before_date,
                status="Confirmed",
                deleted_at__isnull=True,
            ).all()
            for r in rows:
                amt = _q_money(r.total_amount)
                if not _is_refund_voucher(getattr(r, "settlement_type", None), amt):
                    total += _abs_money(amt)

            return_rows = await Receivable.filter(
                tenant_id=tenant_id,
                customer_id=partner_id,
                business_date__lt=before_date,
                deleted_at__isnull=True,
                source_type=RECEIVABLE_SOURCE_SALES_RETURN,
            ).exclude(review_status__in=_EXCLUDED_REVIEW).all()
            total += sum((r.total_amount for r in return_rows), Decimal("0"))
        else:
            rows = await Payment.filter(
                tenant_id=tenant_id,
                supplier_id=partner_id,
                payment_date__lt=before_date,
                status="Confirmed",
                deleted_at__isnull=True,
            ).all()
            for p in rows:
                amt = _q_money(p.total_amount)
                if not _is_refund_voucher(getattr(p, "settlement_type", None), amt):
                    total += _abs_money(amt)

            return_rows = await Payable.filter(
                tenant_id=tenant_id,
                supplier_id=partner_id,
                business_date__lt=before_date,
                deleted_at__isnull=True,
                source_type=PAYABLE_SOURCE_PURCHASE_RETURN,
            ).exclude(review_status__in=_EXCLUDED_REVIEW).all()
            total += sum((p.total_amount for p in return_rows), Decimal("0"))
        return _q_money(total)

    async def _calc_opening_balance(
        self, tenant_id: int, partner_id: int, partner_type: str, start_date: date
    ) -> Decimal:
        debits = await self._sum_debits_before(tenant_id, partner_id, partner_type, start_date)
        credits = await self._sum_credits_before(tenant_id, partner_id, partner_type, start_date)
        return _q_money(debits - credits)

    async def _collect_period_raw_lines(
        self,
        tenant_id: int,
        partner_id: int,
        partner_type: str,
        start_date: date,
        end_date: date,
    ) -> List[Dict[str, Any]]:
        lines: List[Dict[str, Any]] = []
        if partner_type == "Customer":
            receivables = await Receivable.filter(
                tenant_id=tenant_id,
                customer_id=partner_id,
                business_date__gte=start_date,
                business_date__lte=end_date,
                deleted_at__isnull=True,
            ).exclude(review_status__in=_EXCLUDED_REVIEW).order_by("business_date", "id").all()
            receipts = await Receipt.filter(
                tenant_id=tenant_id,
                customer_id=partner_id,
                receipt_date__gte=start_date,
                receipt_date__lte=end_date,
                status="Confirmed",
                deleted_at__isnull=True,
            ).order_by("receipt_date", "id").all()
            for r in receivables:
                amt = _abs_money(r.total_amount)
                is_return = r.source_type == RECEIVABLE_SOURCE_SALES_RETURN
                lines.append({
                    "date": to_api_isoformat(r.business_date),
                    "sort_date": r.business_date,
                    "doc_type": "销售退货" if is_return else "应收单",
                    "doc_code": r.receivable_code,
                    "summary": r.notes or (
                        f"销售退货 {r.source_code or ''}".strip()
                        if is_return
                        else f"应收 {r.source_code or ''}".strip()
                    ),
                    "debit": 0.0 if is_return else float(amt),
                    "credit": float(amt) if is_return else 0.0,
                    "doc_id": r.id,
                    "doc_amount": float(amt),
                })
            for r in receipts:
                amt = _abs_money(r.total_amount)
                is_refund = _is_refund_voucher(getattr(r, "settlement_type", None), r.total_amount)
                lines.append({
                    "date": to_api_isoformat(r.receipt_date),
                    "sort_date": r.receipt_date,
                    "doc_type": "收款退款" if is_refund else "收款单",
                    "doc_code": r.receipt_code,
                    "summary": r.notes or ("收款退款" if is_refund else "收款"),
                    "debit": float(amt) if is_refund else 0.0,
                    "credit": 0.0 if is_refund else float(amt),
                    "doc_id": r.id,
                    "doc_amount": float(amt),
                })
        else:
            payables = await Payable.filter(
                tenant_id=tenant_id,
                supplier_id=partner_id,
                business_date__gte=start_date,
                business_date__lte=end_date,
                deleted_at__isnull=True,
            ).exclude(review_status__in=_EXCLUDED_REVIEW).order_by("business_date", "id").all()
            payments = await Payment.filter(
                tenant_id=tenant_id,
                supplier_id=partner_id,
                payment_date__gte=start_date,
                payment_date__lte=end_date,
                status="Confirmed",
                deleted_at__isnull=True,
            ).order_by("payment_date", "id").all()
            for p in payables:
                amt = _abs_money(p.total_amount)
                is_return = p.source_type == PAYABLE_SOURCE_PURCHASE_RETURN
                line: Dict[str, Any] = {
                    "date": to_api_isoformat(p.business_date),
                    "sort_date": p.business_date,
                    "doc_type": "采购退货" if is_return else "应付单",
                    "doc_code": p.payable_code,
                    "summary": p.notes or (
                        f"采购退货 {p.source_code or ''}".strip()
                        if is_return
                        else f"应付 {p.source_code or ''}".strip()
                    ),
                    "debit": 0.0 if is_return else float(amt),
                    "credit": float(amt) if is_return else 0.0,
                    "doc_id": p.id,
                    "doc_amount": float(amt),
                }
                # 应付 source_type：库内可能是中文常量，也可能是加载 API 的英文码
                src = str(p.source_type or "").strip()
                if p.source_id and src in (
                    PAYABLE_SOURCE_PURCHASE_RECEIPT,
                    "purchase_receipt",
                    "PurchaseReceipt",
                ):
                    line["inbound_detail_doc_type"] = "purchase_receipt"
                    line["inbound_detail_doc_id"] = int(p.source_id)
                elif p.source_id and src in (
                    PAYABLE_SOURCE_OUTSOURCE_RECEIPT,
                    "outsource_material_receipt",
                    "outsource_receipt",
                ):
                    line["inbound_detail_doc_type"] = "outsource_material_receipt"
                    line["inbound_detail_doc_id"] = int(p.source_id)
                lines.append(line)
            for p in payments:
                amt = _abs_money(p.total_amount)
                is_refund = _is_refund_voucher(getattr(p, "settlement_type", None), p.total_amount)
                lines.append({
                    "date": to_api_isoformat(p.payment_date),
                    "sort_date": p.payment_date,
                    "doc_type": "付款退款" if is_refund else "付款单",
                    "doc_code": p.payment_code,
                    "summary": p.notes or ("付款退款" if is_refund else "付款"),
                    "debit": float(amt) if is_refund else 0.0,
                    "credit": 0.0 if is_refund else float(amt),
                    "doc_id": p.id,
                    "doc_amount": float(amt),
                })
        lines.sort(key=lambda x: (x["sort_date"], x["doc_type"], x.get("doc_id", 0)))
        lines = await self._order_lines_by_settlement_hierarchy(
            tenant_id, partner_type, lines
        )
        for ln in lines:
            ln.pop("sort_date", None)
        return lines

    async def _order_lines_by_settlement_hierarchy(
        self,
        tenant_id: int,
        partner_type: str,
        lines: List[Dict[str, Any]],
    ) -> List[Dict[str, Any]]:
        from apps.kuaicaiwu.services.finance_settlement_hierarchy import (
            order_lines_by_settlement_hierarchy,
        )

        if not lines:
            return lines
        if partner_type == "Customer":
            return await order_lines_by_settlement_hierarchy(
                tenant_id,
                lines,
                parent_doc_types={"应收单"},
                child_doc_types={"收款单", "收款退款"},
                rel_source="receivable",
                rel_target="receipt",
                debit_doc_type="Receivable",
                credit_doc_type="Receipt",
            )
        return await order_lines_by_settlement_hierarchy(
            tenant_id,
            lines,
            parent_doc_types={"应付单"},
            child_doc_types={"付款单", "付款退款"},
            rel_source="payable",
            rel_target="payment",
            debit_doc_type="Payable",
            credit_doc_type="Payment",
        )

    async def ensure_statement_line_hierarchy(
        self,
        tenant_id: int,
        partner_type: str,
        opening_balance: Decimal,
        lines: List[Dict[str, Any]],
    ) -> List[Dict[str, Any]]:
        """补齐层级并重算行余额（含本次对账金额）。"""
        if not lines:
            return lines
        working = [_normalize_statement_line_amounts(dict(ln)) for ln in lines]
        if not any(int(ln.get("tree_level") or 0) > 0 for ln in working):
            ordered = await self._order_lines_by_settlement_hierarchy(
                tenant_id, partner_type, working
            )
        else:
            ordered = working
        _, _, _, with_bal = self._apply_running_balance(opening_balance, ordered)
        return with_bal

    def _apply_running_balance(
        self, opening_balance: Decimal, lines: List[Dict[str, Any]]
    ) -> Tuple[Decimal, Decimal, Decimal, List[Dict[str, Any]]]:
        balance = opening_balance
        debit_total = Decimal("0")
        credit_total = Decimal("0")
        result: List[Dict[str, Any]] = []
        for ln in lines:
            debit = _q_money(ln.get("debit", 0))
            credit = _q_money(ln.get("credit", 0))
            debit_total += debit
            credit_total += credit
            balance = _q_money(balance + debit - credit)
            row = dict(ln)
            row["balance"] = float(balance)
            result.append(row)
        return _q_money(debit_total), _q_money(credit_total), balance, result

    async def _load_prior_stated_amounts(
        self,
        tenant_id: int,
        partner_id: int,
        partner_type: str,
        exclude_statement_id: Optional[int] = None,
    ) -> Dict[Tuple[str, int], Decimal]:
        """其它对账单中已累计的对账金额（按单据族+id）。"""
        rows = await PartnerStatement.filter(
            tenant_id=tenant_id,
            partner_id=partner_id,
            partner_type=partner_type,
            deleted_at__isnull=True,
        ).only("id", "transaction_details").all()
        totals: Dict[Tuple[str, int], Decimal] = defaultdict(Decimal)
        for stmt in rows:
            if exclude_statement_id is not None and int(stmt.id) == int(exclude_statement_id):
                continue
            details = stmt.transaction_details or {}
            for ln in details.get("lines") or []:
                key = _line_doc_key(ln.get("doc_type"), ln.get("doc_id"))
                if not key:
                    continue
                totals[key] += _line_statement_amount(ln)
        return dict(totals)

    def _apply_partial_statement_lines(
        self,
        lines: List[Dict[str, Any]],
        prior_map: Dict[Tuple[str, int], Decimal],
        amount_overrides: Optional[Dict[Tuple[str, int], Decimal]] = None,
    ) -> Tuple[List[Dict[str, Any]], int, Decimal]:
        """
        按累计已对金额计算未对余额；默认本次对账金额=未对金额。
        已全部对账的行剔除，并将其借贷净额并入期初调节。
        """
        if not lines:
            return [], 0, Decimal("0.00")
        result: List[Dict[str, Any]] = []
        excluded_count = 0
        excluded_net = Decimal("0.00")
        overrides = amount_overrides or {}
        for ln in lines:
            row = dict(ln)
            doc_amount = _line_doc_amount(row)
            key = _line_doc_key(row.get("doc_type"), row.get("doc_id"))
            prior = prior_map.get(key, Decimal("0")) if key else Decimal("0")
            remaining = _q_money(doc_amount - prior)
            if remaining <= 0:
                excluded_count += 1
                excluded_net += _q_money(row.get("debit", 0)) - _q_money(row.get("credit", 0))
                continue
            statement_amount = remaining
            if key and key in overrides:
                statement_amount = _q_money(overrides[key])
            elif row.get("statement_amount") is not None:
                statement_amount = _q_money(row["statement_amount"])
            if statement_amount <= 0 or statement_amount > remaining:
                code = row.get("doc_code") or row.get("doc_id")
                raise ValidationError(
                    f"单据 {code} 本次对账金额须在 0 与未对金额 {remaining} 之间"
                )
            result.append(
                _apply_statement_amount_to_line(row, statement_amount, doc_amount, prior, remaining)
            )
        return result, excluded_count, _q_money(excluded_net)

    def _build_line_amount_override_map(
        self, line_amounts: Optional[List[Dict[str, Any]]]
    ) -> Dict[Tuple[str, int], Decimal]:
        overrides: Dict[Tuple[str, int], Decimal] = {}
        if not line_amounts:
            return overrides
        for item in line_amounts:
            key = _line_doc_key(item.get("doc_type"), item.get("doc_id"))
            if not key:
                continue
            overrides[key] = _q_money(item.get("statement_amount"))
        return overrides

    def _apply_line_amount_overrides_to_preview_lines(
        self,
        lines: List[Dict[str, Any]],
        prior_map: Dict[Tuple[str, int], Decimal],
        overrides: Dict[Tuple[str, int], Decimal],
    ) -> List[Dict[str, Any]]:
        if not overrides:
            return [_normalize_statement_line_amounts(dict(ln)) for ln in lines]
        rebuilt: List[Dict[str, Any]] = []
        for ln in lines:
            row = dict(ln)
            key = _line_doc_key(row.get("doc_type"), row.get("doc_id"))
            if not key or key not in overrides:
                continue
            doc_amount = _line_doc_amount(row)
            prior = prior_map.get(key, Decimal("0"))
            remaining = _q_money(doc_amount - prior)
            if remaining <= 0:
                continue
            statement_amount = _q_money(overrides[key])
            if statement_amount <= 0 or statement_amount > remaining:
                code = row.get("doc_code") or row.get("doc_id")
                raise ValidationError(
                    f"单据 {code} 本次对账金额须在 0 与未对金额 {remaining} 之间"
                )
            rebuilt.append(
                _apply_statement_amount_to_line(row, statement_amount, doc_amount, prior, remaining)
            )
        return rebuilt

    async def preview_statement(
        self,
        tenant_id: int,
        partner_id: int,
        partner_type: str,
        start_date: date,
        end_date: date,
    ) -> Dict[str, Any]:
        if end_date < start_date:
            raise ValidationError("结束日期不能早于开始日期")
        partner = await self._load_partner_snapshot(tenant_id, partner_id, partner_type)
        opening = await self._calc_opening_balance(tenant_id, partner_id, partner_type, start_date)
        raw_lines = await self._collect_period_raw_lines(
            tenant_id, partner_id, partner_type, start_date, end_date
        )
        prior_map = await self._load_prior_stated_amounts(tenant_id, partner_id, partner_type)
        before_count = len(raw_lines)
        raw_lines, excluded_from_period, excluded_net = self._apply_partial_statement_lines(
            raw_lines, prior_map
        )
        # 已纳入其它对账单的本期发生额并入期初，剩余行续算余额
        opening = _q_money(opening + excluded_net)
        debit_total, credit_total, closing, lines = self._apply_running_balance(opening, raw_lines)
        company_name = await self._get_tenant_company_name(tenant_id)
        balance_label = "应收余额" if partner_type == "Customer" else "应付余额"
        stmt_period = start_date.strftime("%Y-%m")
        existing = await self._get_active_period_statement(
            tenant_id, partner_id, partner_type, stmt_period
        )
        return {
            "partner_id": partner_id,
            "partner_name": partner["name"],
            "partner_type": partner_type,
            "start_date": to_api_isoformat(start_date),
            "end_date": to_api_isoformat(end_date),
            "company_name": company_name,
            "balance_label": balance_label,
            "summary": {
                "opening_balance": float(opening),
                "debit_total": float(debit_total),
                "credit_total": float(credit_total),
                "closing_balance": float(closing),
            },
            "lines": lines,
            "partner_snapshot": partner,
            "excluded_from_period": excluded_from_period,
            # 仅提示：同月可有多张对账单；是否可生成看 lines 是否仍有未纳入单据
            "existing_period_statement_id": existing.id if existing else None,
            "existing_period_statement_code": existing.statement_code if existing else None,
            "existing_period": stmt_period if existing else None,
        }

    async def _get_active_period_statement(
        self,
        tenant_id: int,
        partner_id: int,
        partner_type: str,
        statement_period: str,
    ) -> Optional[PartnerStatement]:
        """同往来同月份最近一张未删除对账单（仅作预览提示，不拦截再生成）。"""
        return (
            await PartnerStatement.filter(
                tenant_id=tenant_id,
                partner_id=partner_id,
                partner_type=partner_type,
                statement_period=statement_period,
                deleted_at__isnull=True,
            )
            .order_by("-id")
            .first()
        )

    def _no_remaining_lines_message(
        self,
        period_label: str,
        *,
        existing_code: Optional[str] = None,
    ) -> str:
        if existing_code:
            return (
                f"该往来单位在 {period_label} 的单据已全部纳入对账单 "
                f"{existing_code}，没有可再生成的明细"
            )
        return f"该往来单位在 {period_label} 没有可纳入对账单的已审核应收/应付或已确认收/付款"

    async def _generate_statement_code(self, tenant_id: int) -> str:
        today = today_site_str()
        return await self.generate_code(tenant_id, "PARTNER_STATEMENT_CODE", prefix=f"DZ{today}")

    async def create_statement(
        self,
        tenant_id: int,
        partner_id: int,
        partner_type: str,
        period: str,
        created_by: int,
        notes: Optional[str] = None,
        attachments: Optional[list] = None,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        line_amounts: Optional[List[Dict[str, Any]]] = None,
    ) -> PartnerStatement:
        if start_date and end_date:
            if end_date < start_date:
                raise ValidationError("结束日期不能早于起始日期")
            period_label = f"{to_api_isoformat(start_date)}~{to_api_isoformat(end_date)}"
            # 列表「对账期间」展示用起始月 YYYY-MM；同月允许多张（按单据去重）
            stmt_period = period or start_date.strftime("%Y-%m")
        else:
            start_date, end_date = period_to_date_range(period)
            period_label = period
            stmt_period = period

        preview = await self.preview_statement(
            tenant_id, partner_id, partner_type, start_date, end_date
        )
        if not (preview.get("lines") or []):
            existing_code = preview.get("existing_period_statement_code")
            raise BusinessLogicError(
                self._no_remaining_lines_message(
                    period_label,
                    existing_code=str(existing_code) if existing_code else None,
                )
            )

        code = await self._generate_statement_code(tenant_id)
        company_name = preview["company_name"]
        summary = dict(preview["summary"])
        lines = list(preview.get("lines") or [])
        overrides = self._build_line_amount_override_map(line_amounts)
        if line_amounts is not None and not overrides:
            raise BusinessLogicError("请至少勾选一行对账明细")
        if overrides:
            prior_map = await self._load_prior_stated_amounts(tenant_id, partner_id, partner_type)
            lines = self._apply_line_amount_overrides_to_preview_lines(lines, prior_map, overrides)
            if not lines:
                raise BusinessLogicError("对账明细不能为空")
            opening = _q_money(summary["opening_balance"])
            debit_total, credit_total, closing, lines = self._apply_running_balance(opening, lines)
            summary = {
                "opening_balance": float(opening),
                "debit_total": float(debit_total),
                "credit_total": float(credit_total),
                "closing_balance": float(closing),
            }
        details = {
            "summary": summary,
            "lines": lines,
            "partner_snapshot": preview["partner_snapshot"],
            "balance_label": preview["balance_label"],
            "period_label": period_label,
        }

        user_info = await self.get_user_info(created_by)
        return await PartnerStatement.create(
            tenant_id=tenant_id,
            statement_code=code,
            partner_id=partner_id,
            partner_name=preview["partner_name"],
            partner_type=partner_type,
            statement_period=stmt_period,
            start_date=start_date,
            end_date=end_date,
            opening_balance=_q_money(summary["opening_balance"]),
            debit_total=_q_money(summary["debit_total"]),
            credit_total=_q_money(summary["credit_total"]),
            closing_balance=_q_money(summary["closing_balance"]),
            status="Draft",
            transaction_details=details,
            company_name=company_name,
            notes=notes,
            attachments=attachments,
            created_by=created_by,
            created_by_name=user_info["name"],
            updated_by=created_by,
            updated_by_name=user_info["name"],
        )

    async def list_statements(
        self,
        tenant_id: int,
        skip: int = 0,
        limit: int = 20,
        partner_type: Optional[str] = None,
        partner_id: Optional[int] = None,
        statement_period: Optional[str] = None,
        status: Optional[str] = None,
        keyword: Optional[str] = None,
        statement_code: Optional[str] = None,
        partner_name: Optional[str] = None,
        created_start_date: Optional[str] = None,
        created_end_date: Optional[str] = None,
        updated_start_date: Optional[str] = None,
        updated_end_date: Optional[str] = None,
        sort_field: Optional[str] = None,
        sort_order: Optional[str] = None,
    ) -> Tuple[List[PartnerStatement], int]:
        from apps.kuaicaiwu.services.finance_list_core import apply_finance_partner_statement_list_filters

        query = PartnerStatement.filter(tenant_id=tenant_id, deleted_at__isnull=True)
        if partner_type:
            query = query.filter(partner_type=partner_type)
        if partner_id:
            query = query.filter(partner_id=partner_id)
        if status:
            query = query.filter(status=status)

        query, order_expr = apply_finance_partner_statement_list_filters(
            query,
            keyword=keyword,
            statement_code=statement_code,
            partner_name=partner_name,
            statement_period=statement_period,
            created_start_date=created_start_date,
            created_end_date=created_end_date,
            updated_start_date=updated_start_date,
            updated_end_date=updated_end_date,
            sort_field=sort_field,
            sort_order=sort_order,
        )

        total = await query.count()
        items = await query.offset(skip).limit(limit).order_by(order_expr, "-id")
        return items, total

    async def get_statement(self, tenant_id: int, statement_id: int) -> PartnerStatement:
        obj = await PartnerStatement.get_or_none(
            tenant_id=tenant_id, id=statement_id, deleted_at__isnull=True
        )
        if not obj:
            raise NotFoundError(f"对账单不存在: {statement_id}")
        return obj

    async def confirm_statement(
        self, tenant_id: int, statement_id: int, user_id: int
    ) -> PartnerStatement:
        obj = await self.get_statement(tenant_id, statement_id)
        if obj.status not in ("Draft", "Disputed"):
            raise BusinessLogicError("仅草稿或异议状态的对账单可确认")
        obj.status = "Confirmed"
        obj.confirmed_at = tortoise_timezone.now()
        obj.confirmed_by = user_id
        operator = await User.filter(id=user_id).first()
        apply_update_audit(obj, operator)
        await obj.save()
        return obj

    async def mark_sent(
        self,
        tenant_id: int,
        statement_id: int,
        user_id: int,
        channel: str,
        notes: Optional[str] = None,
    ) -> PartnerStatement:
        obj = await self.get_statement(tenant_id, statement_id)
        if obj.status != "Confirmed":
            raise BusinessLogicError("请先内部确认对账单后再标记发送")
        allowed = ("export", "print", "wechat_manual", "email_manual", "other")
        if channel not in allowed:
            raise ValidationError(f"发送渠道无效，可选: {', '.join(allowed)}")
        obj.status = "Sent"
        obj.sent_at = tortoise_timezone.now()
        obj.sent_by = user_id
        obj.sent_channel = channel
        if notes:
            obj.notes = (obj.notes or "") + ("\n" if obj.notes else "") + notes
        operator = await User.filter(id=user_id).first()
        apply_update_audit(obj, operator)
        await obj.save()
        return obj

    async def dispute_statement(
        self, tenant_id: int, statement_id: int, reason: str, user_id: Optional[int] = None
    ) -> PartnerStatement:
        obj = await self.get_statement(tenant_id, statement_id)
        if obj.status not in ("Sent", "Confirmed"):
            raise BusinessLogicError("仅已确认或已发送的对账单可记录异议")
        obj.status = "Disputed"
        obj.dispute_reason = reason.strip()
        obj.disputed_at = tortoise_timezone.now()
        if user_id is not None:
            operator = await User.filter(id=user_id).first()
            apply_update_audit(obj, operator)
        await obj.save()
        return obj

    async def update_statement_lines(
        self,
        tenant_id: int,
        statement_id: int,
        line_amounts: List[Dict[str, Any]],
        user_id: int,
    ) -> PartnerStatement:
        obj = await self.get_statement(tenant_id, statement_id)
        if obj.status not in ("Draft", "Disputed"):
            raise BusinessLogicError("仅草稿或有异议的对账单可修改对账金额")
        details = dict(obj.transaction_details or {})
        current_lines = list(details.get("lines") or [])
        if not current_lines:
            raise BusinessLogicError("对账单无明细行")
        prior_map = await self._load_prior_stated_amounts(
            tenant_id,
            obj.partner_id,
            obj.partner_type,
            exclude_statement_id=statement_id,
        )
        overrides = self._build_line_amount_override_map(line_amounts)
        if not overrides:
            raise ValidationError("请提供至少一行对账金额")
        rebuilt = self._apply_line_amount_overrides_to_preview_lines(
            current_lines, prior_map, overrides
        )
        if not rebuilt:
            raise BusinessLogicError("对账明细不能为空")
        opening = _q_money(obj.opening_balance)
        if not any(int(ln.get("tree_level") or 0) > 0 for ln in rebuilt):
            rebuilt = await self._order_lines_by_settlement_hierarchy(
                tenant_id, obj.partner_type, rebuilt
            )
        debit_total, credit_total, closing, lines = self._apply_running_balance(opening, rebuilt)
        summary = {
            "opening_balance": float(opening),
            "debit_total": float(debit_total),
            "credit_total": float(credit_total),
            "closing_balance": float(closing),
        }
        details["lines"] = lines
        details["summary"] = summary
        obj.transaction_details = details
        obj.debit_total = debit_total
        obj.credit_total = credit_total
        obj.closing_balance = closing
        operator = await User.filter(id=user_id).first()
        apply_update_audit(obj, operator)
        await obj.save()
        return obj

    async def delete_statement(self, tenant_id: int, statement_id: int) -> None:
        obj = await self.get_statement(tenant_id, statement_id)
        if obj.status != "Draft":
            raise BusinessLogicError("仅草稿状态的对账单可删除")
        obj.deleted_at = tortoise_timezone.now()
        await obj.save()

    _PURCHASE_RECEIPT_DOC_TYPES = frozenset({
        "purchase_receipt", "采购入库", "purchase_receipts",
    })
    _OUTSOURCE_RECEIPT_DOC_TYPES = frozenset({
        "outsource_material_receipt", "委外收货", "outsource_receipt",
    })

    def _normalize_inbound_doc_type(self, doc_type: str) -> str:
        normalized = (doc_type or "").strip()
        if normalized in self._PURCHASE_RECEIPT_DOC_TYPES:
            return "purchase_receipt"
        if normalized in self._OUTSOURCE_RECEIPT_DOC_TYPES:
            return "outsource_material_receipt"
        raise ValidationError(
            "doc_type 仅支持 purchase_receipt/采购入库 或 outsource_material_receipt/委外收货"
        )

    async def get_statement_line_detail(
        self, tenant_id: int, doc_type: str, doc_id: int
    ) -> Dict[str, Any]:
        """返回采购入库或委外收货明细行（含质检字段），供对账单入库明细弹窗。"""
        kind = self._normalize_inbound_doc_type(doc_type)
        if kind == "purchase_receipt":
            return await self._build_purchase_receipt_line_detail(tenant_id, doc_id)
        return await self._build_outsource_receipt_line_detail(tenant_id, doc_id)

    async def _build_purchase_receipt_line_detail(
        self, tenant_id: int, receipt_id: int
    ) -> Dict[str, Any]:
        from apps.kuaizhizao.models.purchase_receipt import PurchaseReceipt
        from apps.kuaizhizao.models.purchase_receipt_item import PurchaseReceiptItem
        from apps.kuaizhizao.models.incoming_inspection import IncomingInspection

        receipt = await PurchaseReceipt.get_or_none(
            tenant_id=tenant_id, id=receipt_id, deleted_at__isnull=True
        )
        if not receipt:
            raise NotFoundError(f"采购入库单不存在: {receipt_id}")

        items = await PurchaseReceiptItem.filter(
            tenant_id=tenant_id, receipt_id=receipt_id, deleted_at__isnull=True
        ).order_by("id").all()
        inspections = await IncomingInspection.filter(
            tenant_id=tenant_id,
            purchase_receipt_id=receipt_id,
            deleted_at__isnull=True,
        ).all()
        inspection_by_material: Dict[int, IncomingInspection] = {}
        for insp in inspections:
            mid = int(insp.material_id)
            if mid not in inspection_by_material:
                inspection_by_material[mid] = insp

        detail_items: List[Dict[str, Any]] = []
        for item in items:
            insp = inspection_by_material.get(int(item.material_id))
            process_waste, material_waste = _extract_waste_quantities(
                getattr(insp, "other_checks", None) if insp else None
            )
            inspection_date = None
            if insp and insp.inspection_time:
                inspection_date = to_api_isoformat(insp.inspection_time)
            insp_qty = None
            if insp is not None and getattr(insp, "inspection_quantity", None) is not None:
                insp_qty = float(insp.inspection_quantity)
            detail_items.append({
                "material_code": item.material_code,
                "material_name": item.material_name,
                "material_spec": item.material_spec,
                "unit": item.material_unit,
                "quantity": float(item.receipt_quantity or 0),
                "unit_price": float(item.unit_price or 0),
                "amount": float(item.total_amount or 0),
                "inspection_quantity": insp_qty,
                "qualified_quantity": float(
                    (insp.qualified_quantity if insp and insp.qualified_quantity is not None else item.qualified_quantity)
                    or 0
                ),
                "unqualified_quantity": float(
                    (
                        insp.unqualified_quantity
                        if insp and insp.unqualified_quantity is not None
                        else item.unqualified_quantity
                    )
                    or 0
                ),
                "quality_status": (
                    getattr(insp, "quality_status", None) if insp else None
                ) or item.quality_status,
                "inspection_date": inspection_date,
                "inspection_passed": _inspection_passed(insp),
                "defect_reason": getattr(insp, "nonconformance_reason", None) if insp else None,
                "process_waste_qty": process_waste,
                "material_waste_qty": material_waste,
            })

        return {
            "doc_type": "purchase_receipt",
            "doc_id": receipt.id,
            "doc_code": receipt.receipt_code,
            "partner_name": receipt.supplier_name,
            "items": detail_items,
        }

    async def _build_outsource_receipt_line_detail(
        self, tenant_id: int, receipt_id: int
    ) -> Dict[str, Any]:
        from apps.kuaizhizao.models.outsource_work_order import (
            OutsourceMaterialReceipt,
            OutsourceWorkOrder,
        )

        receipt = await OutsourceMaterialReceipt.get_or_none(
            tenant_id=tenant_id, id=receipt_id, deleted_at__isnull=True
        )
        if not receipt:
            raise NotFoundError(f"委外收货单不存在: {receipt_id}")

        work_order = await OutsourceWorkOrder.get_or_none(
            tenant_id=tenant_id,
            id=receipt.outsource_work_order_id,
            deleted_at__isnull=True,
        )
        unit_price = Decimal(str(work_order.unit_price or 0)) if work_order else Decimal("0")
        receipt_qty = Decimal(str(receipt.quantity or 0))
        # 与应付生成一致：按合格数量（无合格则按收货数量）× 委外单价
        settle_qty = Decimal(str(receipt.qualified_quantity or receipt.quantity or 0))
        amount = (settle_qty * unit_price).quantize(_MONEY)
        inspection_date = (
            to_api_isoformat(receipt.received_at) if receipt.received_at else None
        )
        unqualified = Decimal(str(receipt.unqualified_quantity or 0))
        qualified = Decimal(str(receipt.qualified_quantity or 0))
        passed = None
        if receipt.received_at is not None or (qualified + unqualified) > 0:
            passed = unqualified <= 0
        process_waste_raw = receipt.process_waste_qty
        material_waste_raw = receipt.material_waste_qty
        defect_reason = receipt.nonconformance_reason or ((receipt.remarks or "").strip() or None)

        detail_items = [{
            "material_code": work_order.product_code if work_order else "",
            "material_name": work_order.product_name if work_order else "",
            "material_spec": getattr(work_order, "product_spec", None) if work_order else None,
            "unit": receipt.unit,
            "quantity": float(receipt_qty),
            "unit_price": float(unit_price),
            "amount": float(amount),
            "inspection_quantity": float(receipt_qty),
            "qualified_quantity": float(qualified),
            "unqualified_quantity": float(unqualified),
            "quality_status": "合格" if passed else ("不合格" if passed is False else None),
            "inspection_date": inspection_date,
            "inspection_passed": passed,
            "defect_reason": defect_reason,
            "process_waste_qty": float(process_waste_raw) if process_waste_raw is not None else None,
            "material_waste_qty": float(material_waste_raw) if material_waste_raw is not None else None,
        }]

        return {
            "doc_type": "outsource_material_receipt",
            "doc_id": receipt.id,
            "doc_code": receipt.code,
            "partner_name": work_order.supplier_name if work_order else None,
            "items": detail_items,
        }

    async def _statement_export_context(self, obj: PartnerStatement) -> Dict[str, Any]:
        details = dict(obj.transaction_details or {})
        lines = await self.ensure_statement_line_hierarchy(
            obj.tenant_id,
            obj.partner_type,
            _q_money(obj.opening_balance),
            list(details.get("lines") or []),
        )
        return {
            "statement_code": obj.statement_code,
            "company_name": obj.company_name or "本公司",
            "partner_name": obj.partner_name,
            "partner_type": obj.partner_type,
            "period": obj.statement_period,
            "start_date": to_api_isoformat(obj.start_date),
            "end_date": to_api_isoformat(obj.end_date),
            "status": obj.status,
            "balance_label": details.get("balance_label") or (
                "应收余额" if obj.partner_type == "Customer" else "应付余额"
            ),
            "summary": details.get("summary") or {
                "opening_balance": float(obj.opening_balance),
                "debit_total": float(obj.debit_total),
                "credit_total": float(obj.credit_total),
                "closing_balance": float(obj.closing_balance),
            },
            "lines": lines,
            "partner_snapshot": details.get("partner_snapshot") or {},
            "notes": obj.notes,
        }

    @staticmethod
    def _export_doc_type_label(ln: Dict[str, Any]) -> str:
        doc_type = str(ln.get("doc_type") or "")
        if int(ln.get("tree_level") or 0) > 0:
            return f"  └ {doc_type}"
        return doc_type

    async def export_excel(self, obj: PartnerStatement) -> BytesIO:
        ctx = await self._statement_export_context(obj)
        wb = Workbook()
        ws = wb.active
        ws.title = "往来对账单"
        title_font = Font(name="Microsoft YaHei", size=14, bold=True)
        header_font = Font(name="Microsoft YaHei", size=11, bold=True)
        body_font = Font(name="Microsoft YaHei", size=11)

        ws["A1"] = f"{ctx['company_name']} — 往来对账单"
        ws["A1"].font = title_font
        ws.merge_cells("A1:G1")
        ws["A2"] = f"对账单号：{ctx['statement_code']}"
        ws["A3"] = f"往来单位：{ctx['partner_name']}"
        ws["A4"] = f"对账期间：{ctx['start_date']} 至 {ctx['end_date']}"
        snap = ctx.get("partner_snapshot") or {}
        if snap.get("finance_contact_name"):
            ws["A5"] = f"财务联系人：{snap.get('finance_contact_name')} {snap.get('finance_contact_phone') or ''}"

        row = 7
        summary = ctx["summary"]
        ws.cell(row, 1, "期初余额").font = header_font
        ws.cell(row, 2, summary["opening_balance"])
        ws.cell(row, 3, "本期借方").font = header_font
        ws.cell(row, 4, summary["debit_total"])
        ws.cell(row, 5, "本期贷方").font = header_font
        ws.cell(row, 6, summary["credit_total"])
        ws.cell(row, 7, "期末余额").font = header_font
        row += 1
        ws.cell(row, 1, summary["opening_balance"])
        ws.cell(row, 3, summary["debit_total"])
        ws.cell(row, 5, summary["credit_total"])
        ws.cell(row, 7, summary["closing_balance"])

        row += 2
        headers = ["日期", "单据类型", "单号", "摘要", "借方", "贷方", ctx["balance_label"]]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row, col, h)
            c.font = header_font
            c.alignment = Alignment(horizontal="center")
        row += 1
        for ln in ctx["lines"]:
            ws.cell(row, 1, ln.get("date", "")).font = body_font
            ws.cell(row, 2, self._export_doc_type_label(ln)).font = body_font
            ws.cell(row, 3, ln.get("doc_code", "")).font = body_font
            ws.cell(row, 4, ln.get("summary", "")).font = body_font
            ws.cell(row, 5, ln.get("debit", 0)).font = body_font
            ws.cell(row, 6, ln.get("credit", 0)).font = body_font
            ws.cell(row, 7, ln.get("balance", 0)).font = body_font
            row += 1

        row += 1
        ws.cell(row, 1, "请于收到本对账单后 7 个工作日内核对并回复；如有异议请注明。").font = body_font

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        return out

    async def render_html(self, obj: PartnerStatement) -> str:
        ctx = await self._statement_export_context(obj)
        summary = ctx["summary"]
        snap = ctx.get("partner_snapshot") or {}
        contact = ""
        if snap.get("finance_contact_name"):
            contact = f"<p>财务联系人：{snap['finance_contact_name']} {snap.get('finance_contact_phone') or ''}</p>"

        rows_html = ""
        for ln in ctx["lines"]:
            rows_html += f"""
            <tr>
              <td>{ln.get('date', '')}</td>
              <td>{self._export_doc_type_label(ln)}</td>
              <td>{ln.get('doc_code', '')}</td>
              <td>{ln.get('summary', '')}</td>
              <td class="num">{ln.get('debit', 0):,.2f}</td>
              <td class="num">{ln.get('credit', 0):,.2f}</td>
              <td class="num">{ln.get('balance', 0):,.2f}</td>
            </tr>"""

        return f"""<!DOCTYPE html>
<html lang="zh-CN"><head><meta charset="utf-8"/>
<title>往来对账单 {ctx['statement_code']}</title>
<style>
  body {{ font-family: "Microsoft YaHei", sans-serif; font-size: 12px; color: #333; padding: 24px; }}
  h1 {{ font-size: 18px; text-align: center; margin-bottom: 8px; }}
  .meta p {{ margin: 4px 0; }}
  table {{ width: 100%; border-collapse: collapse; margin-top: 16px; }}
  th, td {{ border: 1px solid #ccc; padding: 6px 8px; }}
  th {{ background: #f5f5f5; }}
  .num {{ text-align: right; }}
  .summary {{ margin-top: 12px; }}
  .footer {{ margin-top: 20px; color: #666; font-size: 11px; }}
  @media print {{ body {{ padding: 0; }} }}
</style></head><body>
<h1>{ctx['company_name']} — 往来对账单</h1>
<div class="meta">
  <p>对账单号：{ctx['statement_code']}</p>
  <p>往来单位：{ctx['partner_name']}</p>
  <p>对账期间：{ctx['start_date']} 至 {ctx['end_date']}</p>
  {contact}
</div>
<div class="summary">
  <p>期初余额：<strong>{summary['opening_balance']:,.2f}</strong>
     &nbsp;|&nbsp; 本期借方：<strong>{summary['debit_total']:,.2f}</strong>
     &nbsp;|&nbsp; 本期贷方：<strong>{summary['credit_total']:,.2f}</strong>
     &nbsp;|&nbsp; 期末余额：<strong>{summary['closing_balance']:,.2f}</strong></p>
</div>
<table>
  <thead><tr>
    <th>日期</th><th>单据类型</th><th>单号</th><th>摘要</th>
    <th>借方</th><th>贷方</th><th>{ctx['balance_label']}</th>
  </tr></thead>
  <tbody>{rows_html}</tbody>
</table>
<p class="footer">请于收到本对账单后 7 个工作日内核对并回复；如有异议请注明。</p>
</body></html>"""

    async def export_pdf(self, obj: PartnerStatement) -> Tuple[bytes, str]:
        html = await self.render_html(obj)
        try:
            from apps.kuaizhizao.services.print_service import _html_to_pdf_bytes_playwright_async
            pdf_bytes = await _html_to_pdf_bytes_playwright_async(html)
            return pdf_bytes, "application/pdf"
        except Exception:
            return html.encode("utf-8"), "text/html; charset=utf-8"
