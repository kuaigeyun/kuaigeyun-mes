"""材料验收单 Excel 版式对齐打印模板。"""

from io import BytesIO

from openpyxl import load_workbook

from apps.haoligo.services.finance_acceptance_excel import (
    COPY_BODY_ROWS,
    PAGE_ROWS,
    acceptance_excel_filename,
    build_finance_material_acceptance_xlsx,
)


def _padded_lines(n_filled: int = 1) -> list[dict]:
    lines = []
    for i in range(15):
        if i < n_filled:
            lines.append(
                {
                    "line_no": i + 1,
                    "material_code": "",
                    "product_name_spec": f"物料{i + 1}",
                    "quantity_display": "10",
                    "unit": "kg",
                    "unit_price_display": "1.5",
                    "amount_display": "15.00",
                    "remark": "",
                }
            )
        else:
            lines.append(
                {
                    "line_no": "",
                    "material_code": "",
                    "product_name_spec": "",
                    "quantity_display": "",
                    "unit": "",
                    "unit_price_display": "",
                    "amount_display": "",
                    "remark": "",
                }
            )
    return lines


def _sample_data(**overrides):
    data = {
        "company_name": "无锡好力泵业有限公司",
        "company_address": "江苏省无锡市惠山区阳山镇人民西路",
        "invoice_nos": "26442000004359167806",
        "preparer_name": "张三",
        "verifier_name": "张三",
        "supplier_name": "广东聚石化学股份有限公司",
        "sheet_no": "YS-001",
        "sheet_date": "2026-04-21",
        "total_amount_display": "1705400.00",
        "total_amount_uppercase": "壹佰柒拾万伍仟肆佰元整",
        "remark": "抽检",
        "line_pages": [{"line_items": _padded_lines(1)}],
    }
    data.update(overrides)
    return data


def test_acceptance_excel_dual_copy_layout():
    raw = build_finance_material_acceptance_xlsx(_sample_data())
    ws = load_workbook(BytesIO(raw)).active
    assert ws["A1"].value == "无锡好力泵业有限公司"
    assert "材料验收单" in str(ws["A3"].value).replace(" ", "")
    assert "26442000004359167806" in str(ws["G3"].value)
    assert ws["A6"].value == "序号"
    assert ws["C6"].value == "产品名称及规格"
    assert ws["F6"].value == "不含税单价"
    assert ws["C7"].value == "物料1"
    assert "壹佰柒拾万伍仟肆佰元整" in str(ws["A22"].value)
    assert "1705400.00" in str(ws["F22"].value)
    second_copy_start = COPY_BODY_ROWS + 2
    assert ws.cell(row=second_copy_start, column=1).value == "无锡好力泵业有限公司"
    assert ws.cell(row=second_copy_start + 6, column=3).value == "物料1"
    assert "H47" in (ws.print_area or "").replace("$", "")
    assert int(ws.page_setup.paperSize) == 9


def test_acceptance_excel_paginates_beyond_15_lines():
    page1 = _padded_lines(15)
    page2 = _padded_lines(1)
    raw = build_finance_material_acceptance_xlsx(_sample_data(line_pages=[{"line_items": page1}, {"line_items": page2}]))
    ws = load_workbook(BytesIO(raw)).active
    page2_start = PAGE_ROWS + 1
    assert ws.cell(row=page2_start, column=1).value == "无锡好力泵业有限公司"
    assert ws.cell(row=page2_start + 6, column=3).value == "物料1"


def test_acceptance_excel_filename_strips_invalid_chars():
    assert acceptance_excel_filename("YS/001", 9) == "材料验收单-YS-001.xlsx"
    assert acceptance_excel_filename("", 12) == "材料验收单-12.xlsx"
